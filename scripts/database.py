######connect to remove server via SSH
# then connect to MySQL server
# execute SQL query and return the result
# save the result to .txt and excel file
from sshtunnel import SSHTunnelForwarder
import pymysql
import openpyxl

def dbconnect(sql):
    try:
        with SSHTunnelForwarder(
            ssh_address_or_host=('54.160.73.67', 22),   # ssh 目标服务器 ip 和 port
            ssh_username="bitnami",                           # ssh 目标服务器用户名
            ssh_password=123,                       # ssh 目标服务器用户密码
            ssh_pkey="D:\\www\\SeleniumWWS\\WordPressKeyPair.pem",         # ssh 目标服务器证书
            # ssh_private_key_password=“”                 # ssh 目标服务器证书密码
            remote_bind_address=('127.0.0.1', 3306),        #  mysql 服务 ip 和 port
            local_bind_address=('127.0.0.1', 3306),        # ssh 目标服务器的用于连接 mysql 或 redis 的端口，该 ip 必须为 127.0.0.1
                ) as server:
                    conn = pymysql.connect(
                        host=server.local_bind_host,       # server.local_bind_host 是 参数 local_bind_address 的 ip
                        port=server.local_bind_port,       # server.local_bind_host 是 参数 local_bind_address 的 port
                        user="root",
                        password='hHXMybEYKkf2',
                        db='bitnami_wordpress',
                        charset="utf8"
                    )
                    cursor = conn.cursor()
                    cursor.execute(sql)
                    #tuple type result
                    result = cursor.fetchall()
                    conn.close()
        #print(type(result), result)
    except Exception as e:
        print(e)
        print('error')
    return result

def write_txt(filename,data):
    with open (filename, 'a+') as f:
        f.write(data)
        f.write('\n') 
        f.close()
    #print('file writen successfully')

def write_excel(filename,data):
    wb = openpyxl.load_workbook(filename)
    ws = wb.get_sheet_by_name('Sheet1')
    rownum = ws.max_row
    ws.cell(rownum + 1, 1, data)
    wb.save(filename)


def main():
    try: 
        filename = 'D:\\www\\SeleniumWWS\\dbresult.txt'
        excel_filename = 'D:\\www\\SeleniumWWS\\dbresult.xlsx'
        sql = "select * from wp_posts where post_status = 'publish' and comment_status = 'open';"
        res = dbconnect(sql)
        #print(len(res))
        #print(len(res),type(res),res)
        
        for i in range(0,len(res)):
            print(res[i])
            write_txt(filename,str(res[i]))
            write_excel(excel_filename,str(res[i]))

    except Exception as e:
        print('====scripting error====')
        print(e)

if __name__ == "__main__":
    main()
